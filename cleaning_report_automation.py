# cleaning_report_automation.py
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter

df = pd.read_csv("Chocolate Sales.csv")
df.columns = df.columns.str.strip().str.replace(' ', '_')
df['Amount'] = df['Amount'].astype(str).str.replace('[\$,]', '', regex=True).str.strip()
df['Amount'] = pd.to_numeric(df['Amount'], errors='coerce')
df['Date'] = pd.to_datetime(df['Date'], format='%d-%b-%y', errors='coerce')
df['Boxes_Shipped'] = pd.to_numeric(df['Boxes_Shipped'], errors='coerce')
df = df.dropna(subset=['Amount', 'Date', 'Boxes_Shipped'])
df = df.drop_duplicates()

# Quality metrics
duplicates_before = df.duplicated().sum()
negative_amount = (df['Amount'] <= 0).sum()
negative_boxes = (df['Boxes_Shipped'] <= 0).sum()
outlier_amount = (df['Amount'] > df['Amount'].quantile(0.99)).sum()

# Summaries
sales_by_product = df.groupby('Product')['Amount'].agg(['sum', 'mean', 'count']).round(2)
sales_by_country = df.groupby('Country')['Amount'].agg(['sum', 'mean']).round(2)
monthly_sales = df.groupby(pd.Grouper(key='Date', freq='ME'))['Amount'].sum().reset_index()
monthly_sales.columns = ['Month', 'Total_Sales']

# Visuals
sns.set_style("whitegrid")
plt.rcParams['figure.figsize'] = (10,6)
top_products = sales_by_product.head(10)
plt.barh(top_products.index, top_products['sum'], color='darkorange')
plt.title('Top 10 Products by Total Sales')
plt.tight_layout()
plt.savefig('top_products.png', dpi=100, bbox_inches='tight')
plt.close()

plt.pie(sales_by_country['sum'], labels=sales_by_country.index, autopct='%1.1f%%')
plt.title('Sales Distribution by Country')
plt.savefig('sales_by_country.png', dpi=100, bbox_inches='tight')
plt.close()

plt.plot(monthly_sales['Month'], monthly_sales['Total_Sales'], marker='o', color='green')
plt.title('Monthly Sales Trend')
plt.xticks(rotation=45)
plt.tight_layout()
plt.savefig('monthly_trend.png', dpi=100, bbox_inches='tight')
plt.close()

plt.hist(df['Amount'], bins=30, color='skyblue', edgecolor='black')
plt.title('Distribution of Transaction Amounts')
plt.savefig('amount_distribution.png', dpi=100, bbox_inches='tight')
plt.close()

# Excel report
output_file = "Cleaned_Report.xlsx"
with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    df.to_excel(writer, sheet_name='Cleaned_Data', index=False)
    quality_df = pd.DataFrame({
        'Metric': ['Original Rows', 'Cleaned Rows', 'Duplicates Removed', 'Missing Values (after cleaning)',
                   'Negative Amounts', 'Negative Boxes Shipped', 'Outliers (Amount > 99th percentile)'],
        'Value': [df.shape[0], df.shape[0], duplicates_before, df.isnull().sum().sum(),
                  negative_amount, negative_boxes, outlier_amount]
    })
    quality_df.to_excel(writer, sheet_name='Data_Quality', index=False)
    df.describe(include='all').round(2).to_excel(writer, sheet_name='Summary_Stats')
    sales_by_product.to_excel(writer, sheet_name='Sales_by_Product')
    sales_by_country.to_excel(writer, sheet_name='Sales_by_Country')
    monthly_sales.to_excel(writer, sheet_name='Monthly_Trend', index=False)
    
    # Add images
    ws = writer.book.create_sheet("Charts")
    ws.add_image(Image('top_products.png'), 'A1')
    ws.add_image(Image('sales_by_country.png'), 'A20')
    ws.add_image(Image('monthly_trend.png'), 'K1')
    ws.add_image(Image('amount_distribution.png'), 'K20')
    
    for sheet in writer.book.worksheets:
        for col in sheet.columns:
            max_len = max((len(str(cell.value)) for cell in col), default=0)
            sheet.column_dimensions[get_column_letter(col[0].column)].width = min(max_len+2, 50)

print("✅ Cleaned report generated: Cleaned_Report.xlsx")