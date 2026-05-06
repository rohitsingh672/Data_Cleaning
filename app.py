# app.py – Complete Integrated Dashboard
# Run: streamlit run app.py

import streamlit as st
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from datetime import datetime
from sklearn.linear_model import LinearRegression
from sklearn.ensemble import RandomForestRegressor
from sklearn.metrics import mean_absolute_error, mean_squared_error, r2_score
from sklearn.model_selection import train_test_split
from sklearn.preprocessing import LabelEncoder, StandardScaler
from statsmodels.tsa.holtwinters import ExponentialSmoothing
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
import io
import warnings
warnings.filterwarnings('ignore')

st.set_page_config(page_title="Chocolate Sales – Complete Analytics", layout="wide")
st.title("🍫 Chocolate Sales – Complete Data Analytics Platform")
st.markdown("**Predictive Analytics** + **Data Cleaning & Automated Reporting** in one dashboard.")

# -------------------------------
# 1. LOAD & CLEAN DATA (Shared)
# -------------------------------
@st.cache_data
def load_and_clean_data():
    df = pd.read_csv("Chocolate Sales.csv")
    # Clean Amount
    df['Amount'] = df['Amount'].astype(str).str.replace('[\$,]', '', regex=True).str.strip()
    df['Amount'] = pd.to_numeric(df['Amount'], errors='coerce')
    # Parse Date
    df['Date'] = pd.to_datetime(df['Date'], format='%d-%b-%y', errors='coerce')
    # Drop invalid
    df = df.dropna(subset=['Date', 'Amount'])
    df['Boxes Shipped'] = pd.to_numeric(df['Boxes Shipped'], errors='coerce')
    df = df.dropna(subset=['Boxes Shipped'])
    # Remove duplicates
    df = df.drop_duplicates()
    # Add time features
    df['Year'] = df['Date'].dt.year
    df['Month'] = df['Date'].dt.month
    df['DayOfWeek'] = df['Date'].dt.dayofweek
    df['Quarter'] = df['Date'].dt.quarter
    return df

df = load_and_clean_data()

# -------------------------------
# 2. MAIN TABS
# -------------------------------
tab1, tab2, tab3 = st.tabs(["📈 Predictive Analytics", "🧹 Data Cleaning & Reporting", "📊 Data Overview & EDA"])

# ================= TAB 1: PREDICTIVE ANALYTICS =================
with tab1:
    st.header("Predictive Analytics")
    target_col = st.selectbox("Target variable", ["Amount", "Boxes Shipped"], key="target")
    
    st.subheader("🔮 Time Series Forecasting")
    freq = st.selectbox("Aggregation period", ["Daily", "Weekly", "Monthly"], index=2)
    if freq == "Daily":
        ts_df = df.groupby('Date')[target_col].sum().reset_index()
        ts_df = ts_df.set_index('Date').asfreq('D')
        seasonal = 7
    elif freq == "Weekly":
        ts_df = df.groupby(pd.Grouper(key='Date', freq='W'))[target_col].sum().reset_index()
        ts_df = ts_df.set_index('Date').asfreq('W')
        seasonal = 52
    else:
        ts_df = df.groupby(pd.Grouper(key='Date', freq='ME'))[target_col].sum().reset_index()
        ts_df = ts_df.set_index('Date').asfreq('ME')
        seasonal = 12

    ts_df[target_col] = ts_df[target_col].interpolate(method='linear')
    ts_df = ts_df.dropna()
    st.line_chart(ts_df[target_col], height=250)

    test_size = st.slider("Test proportion (last %)", 0.1, 0.3, 0.2, key="ts_test")
    split_idx = int(len(ts_df) * (1 - test_size))
    train = ts_df.iloc[:split_idx]
    test = ts_df.iloc[split_idx:]

    if len(train) < 2 * seasonal:
        st.warning(f"Less than 2 seasonal cycles ({seasonal}) → using non‑seasonal model.")
        seasonal = None

    try:
        if seasonal:
            model = ExponentialSmoothing(train[target_col], trend='add', seasonal='add', seasonal_periods=seasonal)
        else:
            model = ExponentialSmoothing(train[target_col], trend='add', seasonal=None)
        fitted = model.fit()
        preds = fitted.forecast(len(test))
        mae = mean_absolute_error(test[target_col], preds)
        rmse = np.sqrt(mean_squared_error(test[target_col], preds))
        r2 = r2_score(test[target_col], preds)
        st.success(f"**Metrics:** MAE={mae:.2f}, RMSE={rmse:.2f}, R²={r2:.3f}")

        fig, ax = plt.subplots(figsize=(10,4))
        ax.plot(test.index, test[target_col], label='Actual')
        ax.plot(test.index, preds, label='Predicted', linestyle='--')
        ax.legend()
        st.pyplot(fig)

        periods = st.number_input("Forecast periods", 1, 52, 12)
        future = fitted.forecast(periods)
        st.dataframe(pd.DataFrame({f'Forecasted_{target_col}': future}))
        fig2, ax2 = plt.subplots(figsize=(12,4))
        ax2.plot(ts_df.index, ts_df[target_col], label='Historical')
        ax2.plot(future.index, future, label='Forecast', linestyle='--')
        ax2.legend()
        st.pyplot(fig2)
    except Exception as e:
        st.error(f"Forecast failed: {e}")

    st.subheader("🤖 Regression Modeling")
    cat_cols = ['Sales Person', 'Country', 'Product']
    num_cols = ['Boxes Shipped', 'Year', 'Month', 'DayOfWeek', 'Quarter']
    df_ml = df.copy()
    encoders = {}
    for col in cat_cols:
        le = LabelEncoder()
        df_ml[col+'_enc'] = le.fit_transform(df_ml[col])
        encoders[col] = le
    feature_cols = num_cols + [c+'_enc' for c in cat_cols]
    X = df_ml[feature_cols]
    y = df_ml[target_col]

    test_ratio = st.slider("Test split ratio", 0.1, 0.4, 0.2, key="reg_test")
    X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=test_ratio, random_state=42)
    model_type = st.selectbox("Model", ["Linear Regression", "Random Forest"])
    if model_type == "Linear Regression":
        model = LinearRegression()
    else:
        model = RandomForestRegressor(n_estimators=100, random_state=42)
    scale = st.checkbox("Standardize features")
    if scale:
        scaler = StandardScaler()
        X_train = scaler.fit_transform(X_train)
        X_test = scaler.transform(X_test)
    model.fit(X_train, y_train)
    y_pred = model.predict(X_test)
    st.write(f"**MAE:** {mean_absolute_error(y_test, y_pred):.2f}, **RMSE:** {np.sqrt(mean_squared_error(y_test, y_pred)):.2f}, **R²:** {r2_score(y_test, y_pred):.3f}")

    fig, ax = plt.subplots()
    ax.scatter(y_test, y_pred, alpha=0.5)
    ax.plot([y_test.min(), y_test.max()], [y_test.min(), y_test.max()], 'r--')
    ax.set_xlabel("Actual")
    ax.set_ylabel("Predicted")
    st.pyplot(fig)

    if model_type == "Random Forest" and not scale:
        imp = pd.DataFrame({'Feature': feature_cols, 'Importance': model.feature_importances_}).sort_values('Importance', ascending=False)
        st.bar_chart(imp.set_index('Feature'))

    st.subheader("Manual Prediction")
    with st.form("pred_form"):
        cols = st.columns(3)
        with cols[0]:
            sp = st.selectbox("Sales Person", df['Sales Person'].unique())
            country = st.selectbox("Country", df['Country'].unique())
        with cols[1]:
            product = st.selectbox("Product", df['Product'].unique())
            boxes = st.number_input("Boxes Shipped", 1, 1000, 100)
        with cols[2]:
            pred_date = st.date_input("Date", datetime.now())
        if st.form_submit_button("Predict"):
            inp = {
                'Boxes Shipped': boxes,
                'Year': pred_date.year,
                'Month': pred_date.month,
                'DayOfWeek': pred_date.weekday(),
                'Quarter': (pred_date.month-1)//3 + 1,
                'Sales Person_enc': encoders['Sales Person'].transform([sp])[0],
                'Country_enc': encoders['Country'].transform([country])[0],
                'Product_enc': encoders['Product'].transform([product])[0],
            }
            inp_df = pd.DataFrame([inp])[feature_cols]
            if scale:
                inp_df = scaler.transform(inp_df)
            pred_val = model.predict(inp_df)[0]
            st.success(f"Predicted {target_col}: **${pred_val:,.2f}**" if target_col=="Amount" else f"**{pred_val:,.0f} boxes**")

# ================= TAB 2: DATA CLEANING & REPORTING =================
with tab2:
    st.header("🧹 Data Cleaning & Automated Reporting")
    st.markdown("Generate an **Excel report** with cleaned data, quality metrics, summaries, and charts.")
    
    # Show cleaning summary
    st.subheader("Data Quality Summary")
    col1, col2, col3 = st.columns(3)
    col1.metric("Original Rows", len(df))
    col2.metric("Duplicates Removed", 0)  # already removed
    col3.metric("Missing Values (after cleaning)", df.isnull().sum().sum())
    col1.metric("Negative Amounts", (df['Amount'] <= 0).sum())
    col2.metric("Negative Boxes Shipped", (df['Boxes Shipped'] <= 0).sum())
    col3.metric("Outliers (Amount > 99th %ile)", (df['Amount'] > df['Amount'].quantile(0.99)).sum())
    
    # Button to generate Excel report
    if st.button("📥 Generate Excel Report (Cleaned_Report.xlsx)"):
        with st.spinner("Creating Excel report..."):
            # Prepare data for report
            report_df = df.copy()
            # Quality metrics
            quality_data = {
                'Metric': ['Original Rows (after cleaning)', 'Duplicates Removed', 'Missing Values (after cleaning)',
                           'Negative Amounts', 'Negative Boxes Shipped', 'Outliers (Amount > 99th percentile)'],
                'Value': [len(df), 0, df.isnull().sum().sum(), (df['Amount'] <= 0).sum(),
                          (df['Boxes Shipped'] <= 0).sum(), (df['Amount'] > df['Amount'].quantile(0.99)).sum()]
            }
            quality_df = pd.DataFrame(quality_data)
            # Summary stats
            summary_stats = df.describe(include='all').round(2)
            # Sales by product & country
            sales_by_product = df.groupby('Product')['Amount'].agg(['sum', 'mean', 'count']).round(2).sort_values('sum', ascending=False)
            sales_by_country = df.groupby('Country')['Amount'].agg(['sum', 'mean']).round(2).sort_values('sum', ascending=False)
            # Monthly trend
            monthly_trend = df.groupby(pd.Grouper(key='Date', freq='ME'))['Amount'].sum().reset_index()
            monthly_trend.columns = ['Month', 'Total_Sales']
            
            # Generate images
            fig1, ax1 = plt.subplots()
            top10 = sales_by_product.head(10)
            ax1.barh(top10.index, top10['sum'], color='darkorange')
            ax1.set_title('Top 10 Products')
            plt.tight_layout()
            img1_path = "top_products.png"
            plt.savefig(img1_path, dpi=100, bbox_inches='tight')
            plt.close()
            
            fig2, ax2 = plt.subplots()
            ax2.pie(sales_by_country['sum'], labels=sales_by_country.index, autopct='%1.1f%%')
            ax2.set_title('Sales by Country')
            plt.savefig("sales_by_country.png", dpi=100, bbox_inches='tight')
            plt.close()
            
            fig3, ax3 = plt.subplots()
            ax3.plot(monthly_trend['Month'], monthly_trend['Total_Sales'], marker='o', color='green')
            ax3.set_title('Monthly Sales Trend')
            plt.xticks(rotation=45)
            plt.tight_layout()
            plt.savefig("monthly_trend.png", dpi=100, bbox_inches='tight')
            plt.close()
            
            fig4, ax4 = plt.subplots()
            ax4.hist(df['Amount'], bins=30, color='skyblue', edgecolor='black')
            ax4.set_title('Amount Distribution')
            plt.savefig("amount_distribution.png", dpi=100, bbox_inches='tight')
            plt.close()
            
            # Write to Excel
            output_path = "Cleaned_Report.xlsx"
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                report_df.to_excel(writer, sheet_name='Cleaned_Data', index=False)
                quality_df.to_excel(writer, sheet_name='Data_Quality', index=False)
                summary_stats.to_excel(writer, sheet_name='Summary_Stats')
                sales_by_product.to_excel(writer, sheet_name='Sales_by_Product')
                sales_by_country.to_excel(writer, sheet_name='Sales_by_Country')
                monthly_trend.to_excel(writer, sheet_name='Monthly_Trend', index=False)
                
                # Add images
                ws = writer.book.create_sheet("Charts")
                ws.add_image(XLImage(img1_path), 'A1')
                ws.add_image(XLImage("sales_by_country.png"), 'A20')
                ws.add_image(XLImage("monthly_trend.png"), 'K1')
                ws.add_image(XLImage("amount_distribution.png"), 'K20')
                
                # Auto-width columns
                for sheet in writer.book.worksheets:
                    for col in sheet.columns:
                        max_len = max((len(str(cell.value)) for cell in col if cell.value), default=0)
                        sheet.column_dimensions[get_column_letter(col[0].column)].width = min(max_len+2, 50)
            
            st.success(f"✅ Report saved as `{output_path}`")
            with open(output_path, "rb") as f:
                st.download_button("📥 Download Excel Report", f, file_name="Cleaned_Report.xlsx")
    
    # Preview cleaned data
    st.subheader("Preview of Cleaned Data")
    st.dataframe(df.head(20))

# ================= TAB 3: DATA OVERVIEW & EDA =================
with tab3:
    st.header("Data Overview & Exploratory Analysis")
    st.subheader("Sample Data")
    st.dataframe(df)
    st.subheader("Summary Statistics")
    st.dataframe(df.describe(include='all').round(2))
    
    # Additional plots
    fig, axes = plt.subplots(2, 2, figsize=(12, 8))
    df.groupby('Product')['Amount'].sum().nlargest(10).plot(kind='barh', ax=axes[0,0], color='teal')
    axes[0,0].set_title('Top 10 Products by Sales')
    df.groupby('Country')['Amount'].sum().plot(kind='pie', ax=axes[0,1], autopct='%1.1f%%')
    axes[0,1].set_title('Sales by Country')
    df.groupby(df['Date'].dt.month)['Amount'].mean().plot(kind='line', marker='o', ax=axes[1,0])
    axes[1,0].set_title('Average Monthly Sales')
    sns.boxplot(x='Product', y='Amount', data=df[df['Product'].isin(df['Product'].value_counts().head(10).index)], ax=axes[1,1])
    axes[1,1].set_title('Amount Distribution (Top 10 Products)')
    axes[1,1].tick_params(axis='x', rotation=45)
    plt.tight_layout()
    st.pyplot(fig)